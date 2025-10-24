from logging import Manager

import openpyxl as xl
from openpyxl.chart import BarChart, Reference
def process_xlsx(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(1, sheet.max_row+1):
        for column in range(1, sheet.max_column+1):
            print(f"Row: {row}, column: {column}")
            cell = sheet.cell(row=row, column=1)
    print(cell.value)
    value = Reference(sheet,
              min_row=2,
              max_row=9,
              min_col=1,
              max_col=2)
    chart = BarChart()
    chart.add_data(value)
    sheet.add_chart(chart, 'AM2')
    wb.save(filename)
a = process_xlsx('Mass Manager.xlsx')


